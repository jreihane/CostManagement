'''
Created on Aug 25, 2020

@author: reihane
'''
from openpyxl.utils.cell import column_index_from_string
from openpyxl import load_workbook

class ExcelProcessor(object):
    '''
    classdocs
    '''
    def get_workbook(self):
        self.book = load_workbook(self.path)
    
    def create_sheet(self, sheet_name):
        self.sheet = self.book.create_sheet(sheet_name)
    
    def save_workbook(self):
        self.book.save(self.path)
        
    def clean_worksheet(self):
        for i in range(1, self.sheet.max_row):
            if self.sheet.cell(i, 1).value is None:
                self.book.delete_row()
    
    def insert_row(self, row_num, row):
        self.sheet['A' + str(row_num)] = row[0]
        self.sheet['B' + str(row_num)] = row[2]
        self.sheet['C' + str(row_num)] = float(row[1].strip('"'))
              
    
    def get_current_cell_value(self, cell):
        return str(self.sheet.cell(2, column_index_from_string(cell)).value)    
    
    def update_total_cell(self, row_header, new_value):
        self.sheet[row_header[0] + '2'] = new_value
        self.sheet[row_header[0] + '1'] = row_header[1]
    
    def format_row(self, row_num, row_format):
        
        for l in ['A','B','C']:
            self.sheet[l + str(row_num)].font = row_format[1]
            self.sheet[l + str(row_num)].alignment = row_format[0]
        
    def format_cells(self, col_format, row_header):
        
        self.sheet[row_header[0] + '2'].font = col_format[1]
        self.sheet[row_header[0] + '1'].font = col_format[1]
        self.sheet[row_header[0] + '1'].alignment = col_format[0]
        self.sheet[row_header[0] + '2'].alignment = col_format[0]

    def __init__(self, path):
        '''
        Constructor
        '''
        self.path = path
        self.get_workbook()