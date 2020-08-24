import csv
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from datetime import datetime
from formatter.FormatFactory import FormatFactory

def manageStatement():
    with open('ANZ (16).csv', newline='') as csvfile:
        costs_list = csv.reader(csvfile, delimiter=',', quotechar='|')
        sheet_name = determineSheetName(next(costs_list))
        path = r"CostManagement2020-SecondHalf1.xlsx"
    
        book = load_workbook(path)
        sheet = book.create_sheet(sheet_name)
        
        cleanSheet()
        csvfile.seek(0)

        for row_num, row in enumerate(costs_list):
            if(len(row) > 1):
                amount = float(row[1].strip('"'))
                if(amount < 0):
                    desc = row[2]
                    
                    formatFactory = FormatFactory()
                    try:
                        inserRow(sheet, row_num+1, row, formatFactory.get_formatter(desc), formatFactory.get_header(desc))
                    except:
                        print("Exception!")
                    
        book.save(path)


def determineSheetName(first_cell_value):
    cost_date = datetime.strptime(first_cell_value[0].strip('"'), '%d/%m/%Y')
    
    return cost_date.strftime("%B") + "-" + cost_date.strftime("%G")
    
def cleanSheet():
    return ""


def inserRow(sheet, row_num, row, row_format, row_header):
    
    sheet['A' + str(row_num)] = row[0]
    sheet['B' + str(row_num)] = row[2]
    sheet['C' + str(row_num)] = float(row[1].strip('"'))
    formatRow(sheet, row_num, row_format)
    updateTotalCost(sheet, row[1], row_format, row_header)
          

def updateTotalCost(sheet, row_value, col_format, row_header):
    current_value = str(sheet.cell(2, column_index_from_string(row_header[0])).value)
    if(current_value == "None"):
        current_value = "0"
    
    formatCells(sheet, col_format, row_header)
    sheet[row_header[0] + '2'] = (float(row_value.strip('"')) + float(current_value.strip('"')))
    sheet[row_header[0] + '1'] = row_header[1]


def formatRow(sheet, row_num, row_format):
    
    for l in ['A','B','C']:
        sheet[l + str(row_num)].font = row_format[1]
        sheet[l + str(row_num)].alignment = row_format[0]
    
def formatCells(sheet, col_format, row_header):
    
    sheet[row_header[0] + '2'].font = col_format[1]
    sheet[row_header[0] + '1'].font = col_format[1]
    sheet[row_header[0] + '1'].alignment = col_format[0]
    sheet[row_header[0] + '2'].alignment = col_format[0]
    
    

manageStatement()